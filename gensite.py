# -*- coding: utf-8 -*-
import requests
import json
import jinja2
import os

APIKEY = 'AIzaSyAl3tgYstQjvgM0krud4pz-PxCo-CXK3I0'

pics = {}

import unicodedata
def strip_accents(s):
   return ''.join(c for c in unicodedata.normalize('NFD', s)
                  if unicodedata.category(c) != 'Mn')
def normalize(s):
    return strip_accents(s).replace(' ','').lower() if s else s


from fuzzywuzzy import fuzz
def findPic(nom):
    found = False
    for pic in pics.keys():
        if fuzz.ratio(pic,nom)>90:
            found = True
            break
    return pics[pic] if found else None

import re
def loadImages():
    loadedimages = os.listdir('images/')
    r = requests.get('https://www.googleapis.com/drive/v3/files?q="0BxG9i5BMTacqRl8yMU9SbW5PY0U"+in+parents&key=%s' % APIKEY)

    content = json.loads(r.content)
    if not 'files'  in content:
        print content
        return
    images = content['files']
    for image in images:
        id = re.match(r'photo_([^\.]+).[a-z]+|logo_([^\.]+).[a-z]+',image['name']).groups()
        if id[0]:
            pics[id[0].lower()] = image['name']
        else:
            pics[id[1].lower()] = image['name']


        if image['name'].encode('utf8') in loadedimages:
            continue

        r = requests.get('https://drive.google.com/uc?id='+image['id']+'&export=download')


        open('images/%s' % image['name'],'w').write(r.content)

loadImages()

def parsegdoc(path):
    import requests
    import re
    response = requests.get(path)
    from openpyxl import load_workbook
    from cStringIO import StringIO
    from textwrap import wrap
    def parseSheet(ws):
        fields=[]
        fgroup = None
        for i,f in enumerate(ws.iter_cols(max_row=2)):
            fgroup = normalize(f[0].value) or fgroup
            field = normalize(f[1].value)
            m = re.match(r'([^\(]+)\(([^/)]+)\)',field)
            if m:
                fdesc = {'id':"%s.%s" % (fgroup,m.groups()[0]), 'data':['nom',m.groups()[1]] }
            else:
                fdesc = {'id':"%s.%s" % (fgroup,field) }
            fields.append(fdesc)

        sheetdata = []
        for j,row in enumerate(ws.iter_rows(min_row=3)):
            item = {}
            for i in range(len(fields)):
                field = fields[i]

                if not isinstance(row[i].value,basestring):
                    values = row[i].value
                else:
                    ritems = row[i].value.split('\n')
                    values = []
                    for ritem in ritems:
                        if 'data' in field.keys():
                            m = re.match(r'([^\(]+)\(([^/)]+)\)',ritem)
                            if m:
                                values.append(dict((fk,m.groups()[fi]) for fi,fk in enumerate(field['data'])))
                            else:
                                values.append(ritem) #dict((fk,'PB') for fk in field['data']))
                        else:
                            values.append(ritem)
                    if len(values) == 1:
                        values = values[0]
                item[fields[i]['id']] = values
            sheetdata.append(item)
        return sheetdata

    f = StringIO(response.content)
    wb = load_workbook(f)
    data = {}
    for k in wb.get_sheet_names():
        data[k] = parseSheet(wb[k])
    return data

wb = parsegdoc('https://docs.google.com/spreadsheets/d/1FFz6SUbp1NzpijHxYXBqMkOdjpui6V5fPP5wp4C7CBU/export?format=xlsx&id=1FFz6SUbp1NzpijHxYXBqMkOdjpui6V5fPP5wp4C7CBU')
wsPers = wb[u'Personnalités']
wsMouv = wb['Mouvements']
wsLien = wb['Liens']
wsLegeC = wb[u'LégendeCouleurs']
wsLegeL = wb[u'LégendeLiens']

print wsLien

couleur_autres = '#414141'
personnes = {}
mouvements = {}
liens = []
id = 1


def getId(n):
    return n.lower().replace('-','').replace(' ','')

elements = {'nodes':[],'edges':[]}

nodeWeights = {}
for lien in wsLien:
    nodeWeights[lien['noms.nom1']] = nodeWeights.get(lien['noms.nom1'],0) + 1
    nodeWeights[lien['noms.nom2']] = nodeWeights.get(lien['noms.nom2'],0) + 1
    elements['edges'].append({'data':{'source':lien['noms.nom1'],'target':lien['noms.nom2'],'label':lien['description.desccarte'] or ''}})

categories = []
nodesimages = []
cat_couleurs = {}
for leg in wsLegeC:
    cat_couleurs[leg['noms.nomgroupe']] = leg['noms.codecouleur']
    categories.append(dict(id=strip_accents(leg['noms.nomgroupe']),nom=leg['noms.nomgroupe'],couleur=leg['noms.codecouleur'],
                            libelle=leg['noms.nomcomplet'],desc=leg['description.descgroupe']))

for pers in wsPers:
    idname = getId(pers['intro.nom'])
    node = {'data':
                {'id':pers['intro.nom'],
                 'shortId': normalize(pers['intro.nom']),
                 'label':pers['intro.nom'],
                 'type':'personne',
                 'cat':pers['intro.categorie'],
                 'couleurfiche': cat_couleurs.get(pers['intro.categorie'],couleur_autres),
                 'poids':nodeWeights.get(pers['intro.nom'],0)
                }}
    pic = findPic(idname)
    if pic:
        node['data'].update({'haspic':1})
        node['data']['pic'] = 'images/'+pic
        nodesimages.append(dict(id=pers['intro.nom'],image='images/'+pic))
    else:
        node['data']['pic'] = 'images/vide.png'
    elements['nodes'].append(node)

for mouv in wsMouv:
    idname = getId(mouv['intro.nom'])
    node = {'data':
                {'id':mouv['intro.nom'],
                 'shortId': normalize(mouv['intro.nom']),
                 'label':mouv['intro.nom'],
                 'type':'mouvement',
                 'cat':mouv['intro.categorie'],
                 'couleurfiche': cat_couleurs.get(mouv['intro.categorie'],couleur_autres),
                 'poids':nodeWeights.get(mouv['intro.nom'],0)
                }}
    pic = findPic(idname)
    if pic:
        node['data'].update({'haspic':1})
        node['data']['pic'] = 'images/'+pic
        nodesimages.append(dict(id=mouv['intro.nom'],image='images/'+pic))
    else:
        node['data']['pic'] = 'images/vide.png'

    elements['nodes'].append(node)



import json
open('reseau/reseau.json','w').write(json.dumps({'elements':elements}, indent=4, separators=(',', ': ')))


def renderTemplate(tpl_path, **context):
    path, filename = os.path.split(tpl_path)
    return jinja2.Environment(
        loader=jinja2.FileSystemLoader(path or './')
    ).get_template(filename).render(**context)




from jinja2 import Environment, PackageLoader, select_autoescape,FileSystemLoader
env = Environment(
    loader=FileSystemLoader('./templates'),
    autoescape=select_autoescape(['html', 'xml'])
)
for node in elements['nodes']:
    open('fiches/%s.html' % node['data']['shortId'],'w').write(env.get_template('fichetmpl.html').render(**node['data']).encode('utf-8'))

open('reseau/reseau.cycss','w').write(env.get_template('reseautmpl.cycss').render(categories=categories,nodesimages=nodesimages).encode('utf-8'))
open('js/lepenscope.js','w').write(env.get_template('lepenscopetmpl.js').render(categories=categories).encode('utf-8'))
open('index.html','w').write(env.get_template('bmdindextmpl.html').render(categories=categories).encode('utf-8'))
